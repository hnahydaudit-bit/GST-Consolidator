import streamlit as st
import json
import pandas as pd
from collections import defaultdict
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment

st.set_page_config(page_title="GSTR-1 Consolidator", layout="wide")

# Custom CSS for Light Blue and White Download Button
st.markdown("""
    <style>
    div.stDownloadButton > button {
        background-color: #D6EAF8 !important;
        color: #000000 !important;
        border: 1px solid #AED6F1 !important;
        border-radius: 5px;
        padding: 0.5rem 1rem;
        font-family: 'Cambria';
    }
    div.stDownloadButton > button:hover {
        background-color: #FFFFFF !important;
        border: 1px solid #3498DB !important;
        color: #3498DB !important;
    }
    </style>
    """, unsafe_allow_ diethyl=True)

st.title("GSTR-1 Month-wise Consolidation")

# Month Labels with Years for FY 2024-25
MONTH_LABELS = [
    "Apr-24", "May-24", "Jun-24", "Jul-24", "Aug-24", "Sep-24", 
    "Oct-24", "Nov-24", "Dec-24", "Jan-25", "Feb-25", "Mar-25"
]

# Mapping portal 'fp' (MMYYYY) to our display labels
FP_MAP = {
    "042024": "Apr-24", "052024": "May-24", "062024": "Jun-24", 
    "072024": "Jul-24", "082024": "Aug-24", "092024": "Sep-24",
    "102024": "Oct-24", "112024": "Nov-24", "122024": "Dec-24", 
    "012025": "Jan-25", "022025": "Feb-25", "032025": "Mar-25"
}

TABLES = [
    ("4", "B2B Invoices - 4A, 4B, 4C, 6B, 6C", "b2b"),
    ("5A", "B2C Invoices - 5A, 5B (Large)", "b2cl"),
    ("7", "B2C Invoices 7 - B2C (Others)", "b2cs"),
    ("6A", "Exports Invoices - 6A", "exp"),
    ("8", "Nil rated, exempted and non GST outward supplies - 8", "nil"),
    ("9B-R", "Credit/Debit Notes (Registered) - 9B", "cdnr"),
    ("9B-U", "Credit/Debit Notes (Unregistered) - 9B", "cdnur"),
    ("11A", "Tax Liability (Advances Received) - 11A", "at"),
    ("11B", "Adjustment of Advances - 11B", "txpd")
]

summary_data = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

uploaded_files = st.file_uploader("Upload Monthly GSTR-1 JSON files", type="json", accept_multiple_files=True)

if uploaded_files:
    for file in uploaded_files:
        try:
            data = json.load(file)
            fp = data.get("fp", "")
            if not fp: continue
            
            m_label = FP_MAP.get(fp)
            if not m_label: continue

            for _, _, key in TABLES:
                section = data.get(key, [])
                if not section: continue

                if key == "nil":
                    inv_list = section.get("inv", []) if isinstance(section, dict) else section
                    for item in inv_list:
                        summary_data[key]["Nil-rated Supply"][m_label] += item.get("nil_amt", 0)
                        summary_data[key]["Exempt Supply"][m_label] += item.get("expt_amt", 0)
                        summary_data[key]["Non-GST Supply"][m_label] += item.get("ngsup_amt", 0)
                elif key == "b2cs":
                    for item in section:
                        summary_data[key]["Taxable Value"][m_label] += item.get("txval", 0)
                        summary_data[key]["IGST"][m_label] += item.get("iamt", 0)
                        summary_data[key]["CGST"][m_label] += item.get("camt", 0)
                        summary_data[key]["SGST"][m_label] += item.get("samt", 0)
                        summary_data[key]["Cess"][m_label] += item.get("csamt", 0)
                elif key in ["at", "txpd"]:
                    for entry in section:
                        for itm in entry.get("itms", []):
                            summary_data[key]["Taxable Value"][m_label] += itm.get("ad_amt", 0)
                            summary_data[key]["IGST"][m_label] += itm.get("iamt", 0)
                            summary_data[key]["CGST"][m_label] += itm.get("camt", 0)
                            summary_data[key]["SGST"][m_label] += itm.get("samt", 0)
                            summary_data[key]["Cess"][m_label] += itm.get("csamt", 0)
                else:
                    for entry in section:
                        docs = entry.get('inv', []) if 'inv' in entry else (entry.get('nt', []) if 'nt' in entry else [entry])
                        for doc in docs:
                            for item in doc.get('itms', []):
                                det = item.get('itm_det', item)
                                summary_data[key]["Taxable Value"][m_label] += det.get("txval", 0)
                                summary_data[key]["IGST"][m_label] += det.get("iamt", 0)
                                summary_data[key]["CGST"][m_label] += det.get("camt", 0)
                                summary_data[key]["SGST"][m_label] += det.get("samt", 0)
                                summary_data[key]["Cess"][m_label] += det.get("csamt", 0)
        except Exception as e:
            st.error(f"Error in {file.name}: {str(e)}")

    if st.button("Generate Consolidated Excel"):
        final_rows = []
        table_name_row_indices = []
        
        for _, tbl_name, key in TABLES:
            table_name_row_indices.append(len(final_rows) + 2) 
            final_rows.append({"Particulars": f"GSTR-1 Summary calculated by Govt. Portal: {tbl_name}"})
            
            tax_rows = ["Nil-rated Supply", "Exempt Supply", "Non-GST Supply"] if key == "nil" else ["Taxable Value", "IGST", "CGST", "SGST", "Cess"]
            for tax in tax_rows:
                row = {"Particulars": tax}
                total = 0
                for m in MONTH_LABELS:
                    val = summary_data[key][tax][m]
                    row[m] = val
                    total += val
                row["Total"] = total
                final_rows.append(row)
            final_rows.append({"Particulars": ""})

        df = pd.DataFrame(final_rows)
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="GSTR-1 Summary")
            ws = writer.sheets["GSTR-1 Summary"]
            
            cambria_normal = Font(name="Cambria", size=11, bold=False)
            cambria_bold = Font(name="Cambria", size=11, bold=True)
            custom_blue = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
            acc_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = cambria_normal

            for cell in ws[1]:
                cell.font = cambria_bold

            for r_idx in table_name_row_indices:
                for c_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.fill = custom_blue
                    cell.font = cambria_bold

            for row in range(2, ws.max_row + 1):
                for col in range(2, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = acc_format

            ws.column_dimensions['A'].width = 65

        st.success("Consolidation successful!")
        st.download_button("Download Consolidated Report", output.getvalue(), "GSTR1_Consolidated_Report.xlsx")
